from flask import Blueprint,render_template,request,Flask,session,redirect,url_for
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from flask_mysqldb import MySQL
import MySQLdb.cursors


wishlist_blueprint= Blueprint('wishlist_blueprint', __name__)
app = Flask(__name__)
mysql = MySQL(app)

# @wishlist_blueprint.route('/addproduct_wishlist', methods=['GET','POST'])
# def addproduct_wishlist():
#     if request.method == "POST":
#         data = request.json
#         if data and "user_id" in data and "product_id" in data:
#             user_id = data["user_id"]
#             product_id = data["product_id"]
#             cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
#             cursor.execute(
#                 "SELECT * FROM user WHERE user_id = %s",
#                 (user_id,),
#             )
#             account = cursor.fetchone()
#             if not account:
#                 return {
#                     "status": "FAILURE",
#                     "message": "user_id does not exists",
#                     "data": "",
#                     "traceback": "",
#                 }
#             cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
#             cursor.execute(
#                 "SELECT * FROM product WHERE product_id = %s",
#                 (product_id,),
#             )
#             account = cursor.fetchone()
#             if not account:
#                 return {
#                     "status": "FAILURE",
#                     "message": " product_id does not exists",
#                     "data": "",
#                     "traceback": "",
#                 }
#             cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
#             cursor.execute(
#                 "SELECT * FROM wishlist WHERE user_id = %s AND product_id=%s",
#                 (user_id, product_id),
#             )
#             account = cursor.fetchone()
#             if account:
#                 return {
#                     "status": "FAILURE",
#                     "message": "user_id and product_id combination already exists",
#                     "data": "",
#                     "traceback": "",
#                 }

#             else:
#                 cursor.execute(
#                     "INSERT INTO wishlist(user_id,product_id) VALUES (%s,%s)",
#                     (user_id, product_id),
#                 )
#             mysql.connection.commit()
#             return {
#                 "status": "SUCESS",
#                 "message": "SUCESSFULLY added to wishlist",
#                 "data": "",
#                 "traceback": "",
#             }
#     return ""
    
    
    
    
@wishlist_blueprint.route("/addproduct_wishlist/<int:product_id>", methods=["GET"])
def addproduct_wishlist(product_id):
    if "logged_in" not in session or not session["logged_in"] or "usertype" not in session or session['usertype']!='user':
        print(session)
        return redirect(url_for("login_blueprint.login"))
    else:
        if product_id:
            user_id = session["user_id"]
            cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute(
                "SELECT * FROM wishlist WHERE user_id = %s AND product_id=%s",
                (user_id, product_id),
            )
            account = cursor.fetchone()
            if account:
                message = "product already exists in your wishlist"
                alert_class = "warning"
            else:
                cursor.execute(
                    "INSERT INTO wishlist (user_id,product_id) VALUES (%s,%s)",
                    (user_id, product_id),
                )
                mysql.connection.commit()
                message = "product successfully added to wishlist"
                alert_class = "success"
                cursor.close()
            return redirect(
                url_for(
                    "wishlist_blueprint.view_wishlist", message=message, alert_class=alert_class
                )
            )
    return redirect(url_for("wishlist_blueprint.view_wishlist",message=message, alert_class=alert_class))
   
    
@wishlist_blueprint.route("/view_wishlist", methods=["GET"])
def view_wishlist():
    if "logged_in" not in session or not session["logged_in"] or "usertype" not in session or session['usertype']!='user':
        return redirect("login")
    else:
        user_id = session["user_id"]
        message = request.args.get("message",'')
        alert_class=request.args.get('alert_class')
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            f"SELECT w.*,p.image,p.name,p.amount FROM wishlist AS w JOIN product as p ON p.product_id=w.product_id WHERE w.user_id={user_id}"
        )
        wishlist_items = cursor.fetchall()
        cursor.close()
        # print(cart_items)
        # print(user_id)

    return render_template(
        "wishlist.html",message=message,alert_class=alert_class,
        wishlist_items=wishlist_items,
    )
  
    
@wishlist_blueprint.route("/delete_wishlist", methods=["POST"])
def delete_wishlist():
    if "logged_in" not in session or not session["logged_in"] or "usertype" not in session or session['usertype']!='user':
        return redirect("login")
    else:
        user_id = session["user_id"]
        product_id = request.form.get("product_id")
        cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            f" DELETE FROM wishlist WHERE user_id={user_id} AND product_id={product_id}"
        )
        mysql.connection.commit()
        cursor.close()
        message="Product successfully removed from your wishlist"
        alert_class='success'
        return redirect(url_for("wishlist_blueprint.view_wishlist",message=message, alert_class=alert_class))

    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    
    # if request.method=='POST':
    #     details=pd.read_excel('open_ecommerce.xlsx',sheet_name='wishlist')
    #     user_id=request.get_json()['user_id']
    #     product_id=request.get_json()['product_id']
        
    #     user_data=pd.read_excel('open_ecommerce.xlsx',sheet_name='user_data')
    #     if user_id not in user_data.index:
    #         return {
    #         'status':'failure',
    #         'message':'Invalid userid',
    #         'data': {},
    #         'traceback':''
    #         }
    #     product_data=pd.read_excel('open_ecommerce.xlsx',sheet_name='products')
    #     if product_id not in product_data.index:
    #         return {
    #         'status':'failure',
    #         'message':'Invalid Productid',
    #         'data': {},
    #         'traceback':''
    #         }
        
    #     wishlist_data=pd.read_excel('open_ecommerce.xlsx',sheet_name='wishlist')
    #     if ((details['user_id']==user_id) & (details['product_id']==product_id)).any():
    #         return {'status':'failure','message':'The given Userid and Productid combination already exists','data':'','traceback':''}

    #     wb=load_workbook('open_ecommerce.xlsx')
    #     ws=wb['wishlist']
    #     wishlist_id_values=[int(row[0]) for row in ws.iter_rows(min_row=2,values_only=True)]
    #     if wishlist_id_values:
    #         wishlist_id=max(wishlist_id_values) + 1
    #     else:
    #         wishlist_id=1
    #     ws.append([wishlist_id,user_id,product_id])
    #     wb.save('open_ecommerce.xlsx')
        
    #     return {
    #         'status':'sucess',
    #         'message':'Product has been added to the wishlist successfully',
    #         'data': {'wishlist_id':int(wishlist_id)},
    #         'traceback':''
    #         }
    # return render_template('wishlist.html')